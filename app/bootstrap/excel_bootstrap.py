from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, timezone
from hashlib import sha1
from pathlib import Path
import re
import sqlite3
from typing import Any

from openpyxl import load_workbook

from app.db import ComponentRepository
from app.knowledge.ontology_store import OntologyEdge, OntologyStore


@dataclass(slots=True)
class ExcelBootstrapResult:
    sheet_count: int
    total_rows: int
    raw_tables: list[str]
    component_count: int
    inventory_count: int


@dataclass(slots=True)
class SheetColumnMap:
    part_number: str | None
    purchase_id: str | None
    name: str | None
    vendor: str | None
    description: str | None
    spec: str | None
    package: str | None
    voltage: str | None
    current: str | None
    purchase_date: str | None
    quantity: str | None
    location: str | None


def bootstrap_from_excel(
    *,
    excel_path: str | Path,
    repo: ComponentRepository,
    ontology_store: OntologyStore,
) -> ExcelBootstrapResult:
    path = Path(excel_path)
    if not path.exists():
        raise FileNotFoundError(f"excel file not found: {path}")

    workbook = load_workbook(path, read_only=True, data_only=True)
    raw_tables: list[str] = []
    total_rows = 0
    components_seen: set[str] = set()
    inventory_seen: set[str] = set()

    ontology_store.nodes.clear()
    ontology_store.edges.clear()
    edge_seen: set[tuple[str, str, str]] = set()

    workbook_node_id = f"workbook:{_safe_token(path.stem)}"
    ontology_store.upsert_node(
        node_id=workbook_node_id,
        name=path.name,
        node_type="Workbook",
        aliases=[path.stem],
        attributes={"path": str(path)},
    )

    with repo.connection() as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS excel_sheet_registry (
                sheet_name TEXT PRIMARY KEY,
                table_name TEXT NOT NULL,
                row_count INTEGER NOT NULL,
                imported_at TEXT NOT NULL
            )
            """
        )
        conn.execute("DELETE FROM excel_sheet_registry")

        for idx, sheet_name in enumerate(workbook.sheetnames, start=1):
            ws = workbook[sheet_name]
            headers = _build_headers(ws)
            rows = _extract_rows(ws, headers)
            if not rows:
                continue

            total_rows += len(rows)
            table_name = _build_table_name(sheet_name=sheet_name, ordinal=idx)
            _rebuild_raw_table(conn=conn, table_name=table_name, headers=headers, rows=rows)
            raw_tables.append(table_name)
            conn.execute(
                "INSERT INTO excel_sheet_registry (sheet_name, table_name, row_count, imported_at) VALUES (?, ?, ?, ?)",
                (
                    sheet_name,
                    table_name,
                    len(rows),
                    datetime.now(timezone.utc).isoformat(),
                ),
            )

            mapping = _build_column_map(headers)
            sheet_node_id, column_node_ids = _build_sheet_column_graph(
                ontology_store=ontology_store,
                edge_seen=edge_seen,
                sheet_name=sheet_name,
                headers=headers,
                table_name=table_name,
                row_count=len(rows),
            )

            _add_edge(
                ontology_store=ontology_store,
                edge_seen=edge_seen,
                source=workbook_node_id,
                target=sheet_node_id,
                relation="contains_sheet",
            )

            for row_index, row in enumerate(rows, start=1):
                part_number = _upsert_canonical_row(
                    conn=conn,
                    row=row,
                    sheet_name=sheet_name,
                    mapping=mapping,
                )
                if part_number:
                    components_seen.add(part_number)
                    inventory_seen.add(part_number)

                _build_record_graph(
                    ontology_store=ontology_store,
                    edge_seen=edge_seen,
                    sheet_name=sheet_name,
                    sheet_node_id=sheet_node_id,
                    row=row,
                    row_index=row_index,
                    column_node_ids=column_node_ids,
                    mapping=mapping,
                    part_number=part_number,
                )

    workbook.close()
    ontology_store.save()

    return ExcelBootstrapResult(
        sheet_count=len(workbook.sheetnames),
        total_rows=total_rows,
        raw_tables=raw_tables,
        component_count=len(components_seen),
        inventory_count=len(inventory_seen),
    )


def _build_headers(ws: Any) -> list[str]:
    row1 = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ()))
    row2 = list(next(ws.iter_rows(min_row=2, max_row=2, values_only=True), ()))
    width = max(ws.max_column or 0, len(row1), len(row2))

    headers: list[str] = []
    current_group = ""
    for idx in range(width):
        top = _clean_text(row1[idx]) if idx < len(row1) else ""
        sub = _clean_text(row2[idx]) if idx < len(row2) else ""

        if top:
            current_group = top

        group = top or current_group
        if group and sub and sub != group:
            name = f"{group}_{sub}"
        elif group:
            name = group
        elif sub:
            name = sub
        else:
            name = f"列{idx + 1}"

        headers.append(name)

    return _dedupe_headers(headers)


def _dedupe_headers(headers: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    out: list[str] = []
    for header in headers:
        base = header.strip() or "未命名列"
        count = seen.get(base, 0) + 1
        seen[base] = count
        out.append(base if count == 1 else f"{base}_{count}")
    return out


def _extract_rows(ws: Any, headers: list[str]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    width = len(headers)
    for source_row in ws.iter_rows(min_row=3, values_only=True):
        values = list(source_row)
        if len(values) < width:
            values.extend([None] * (width - len(values)))
        else:
            values = values[:width]

        normalized = [_normalize_cell(v) for v in values]
        if not any(v not in (None, "") for v in normalized):
            continue

        rows.append(dict(zip(headers, normalized, strict=False)))
    return rows


def _build_table_name(*, sheet_name: str, ordinal: int) -> str:
    ascii_name = re.sub(r"[^0-9a-zA-Z]+", "_", sheet_name).strip("_").lower()
    if not ascii_name:
        digest = sha1(sheet_name.encode("utf-8")).hexdigest()[:8]
        ascii_name = f"sheet_{digest}"
    return f"excel_{ordinal:02d}_{ascii_name}"


def _rebuild_raw_table(
    *,
    conn: sqlite3.Connection,
    table_name: str,
    headers: list[str],
    rows: list[dict[str, Any]],
) -> None:
    types = {header: _infer_sqlite_type([row.get(header) for row in rows]) for header in headers}

    conn.execute(f"DROP TABLE IF EXISTS {_q(table_name)}")

    col_defs = [f'{_q("__excel_row_id")} INTEGER PRIMARY KEY AUTOINCREMENT']
    col_defs.extend(f"{_q(h)} {types[h]}" for h in headers)
    conn.execute(f"CREATE TABLE {_q(table_name)} ({', '.join(col_defs)})")

    if not rows:
        return

    cols_sql = ", ".join(_q(h) for h in headers)
    placeholders = ", ".join(["?"] * len(headers))
    insert_sql = f"INSERT INTO {_q(table_name)} ({cols_sql}) VALUES ({placeholders})"
    payload = [tuple(row.get(h) for h in headers) for row in rows]
    conn.executemany(insert_sql, payload)


def _infer_sqlite_type(values: list[Any]) -> str:
    has_text = False
    has_float = False
    has_int = False

    for value in values:
        if value is None or value == "":
            continue

        if isinstance(value, bool):
            has_int = True
            continue

        if isinstance(value, int):
            has_int = True
            continue

        if isinstance(value, float):
            has_float = True
            continue

        if isinstance(value, (datetime, date)):
            has_text = True
            continue

        if isinstance(value, str):
            parsed = _try_parse_number(value)
            if parsed is None:
                has_text = True
            elif float(parsed).is_integer():
                has_int = True
            else:
                has_float = True
            continue

        has_text = True

    if has_text:
        return "TEXT"
    if has_float:
        return "REAL"
    if has_int:
        return "INTEGER"
    return "TEXT"


def _try_parse_number(value: str) -> float | None:
    candidate = value.strip().replace(",", "")
    if not candidate:
        return None

    if re.fullmatch(r"-?\d+", candidate):
        return float(candidate)
    if re.fullmatch(r"-?\d+\.\d+", candidate):
        return float(candidate)
    return None


def _build_column_map(headers: list[str]) -> SheetColumnMap:
    return SheetColumnMap(
        part_number=_find_first_column(headers, ["物料编码", "料号", "型号", "part", "part_number"]),
        purchase_id=_find_first_column(headers, ["采购单号", "单号"]),
        name=_find_first_column(headers, ["名称", "name"]),
        vendor=_find_first_column(headers, ["厂商", "品牌", "vendor"]),
        description=_find_first_column(headers, ["描述", "description"]),
        spec=_find_first_column(headers, ["规格", "参数", "spec"]),
        package=_find_first_column(headers, ["封装", "package"]),
        voltage=_find_first_column(headers, ["耐压", "电压", "voltage"]),
        current=_find_first_column(headers, ["电流", "额定电流", "current"]),
        purchase_date=_find_first_column(headers, ["购买日期", "采购日期", "date"]),
        quantity=_find_first_column(headers, ["在库数量", "库存数量", "库存信息", "在库信息", "quantity"]),
        location=_find_first_column(headers, ["位置", "库位", "location"]),
    )


def _find_first_column(headers: list[str], keywords: list[str]) -> str | None:
    lowered = [(h, h.lower()) for h in headers]
    for keyword in keywords:
        target = keyword.lower()
        for original, low in lowered:
            if target in low:
                return original
    return None


def _upsert_canonical_row(
    *,
    conn: sqlite3.Connection,
    row: dict[str, Any],
    sheet_name: str,
    mapping: SheetColumnMap,
) -> str | None:
    purchase_id = _to_text(row.get(mapping.purchase_id)) if mapping.purchase_id else ""
    part_number = _to_text(row.get(mapping.part_number)) if mapping.part_number else ""
    name = _to_text(row.get(mapping.name)) if mapping.name else ""
    vendor = _to_text(row.get(mapping.vendor)) if mapping.vendor else ""
    description = _to_text(row.get(mapping.description)) if mapping.description else ""
    spec = _to_text(row.get(mapping.spec)) if mapping.spec else ""
    package = _to_text(row.get(mapping.package)) if mapping.package else ""
    voltage = _to_text(row.get(mapping.voltage)) if mapping.voltage else ""
    current = _to_text(row.get(mapping.current)) if mapping.current else ""
    purchase_date = _to_text(row.get(mapping.purchase_date)) if mapping.purchase_date else ""
    location = _to_text(row.get(mapping.location)) if mapping.location else ""
    quantity = _to_int(row.get(mapping.quantity)) if mapping.quantity else None

    if not part_number:
        part_number = purchase_id or _make_fallback_part_number(sheet_name=sheet_name, name=name, row=row)

    if not part_number:
        return None

    if not name:
        name = part_number

    desc_fragments = [
        f"purchase_id={purchase_id}" if purchase_id else "",
        f"vendor={vendor}" if vendor else "",
        description,
        f"spec={spec}" if spec else "",
    ]
    merged_description = "; ".join([frag for frag in desc_fragments if frag])

    now = datetime.now(timezone.utc).isoformat()
    conn.execute(
        """
        INSERT INTO components
        (part_number, name, category, package, voltage, current, manufacturer, purchase_date, purchase_id, description, created_at, updated_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(part_number) DO UPDATE SET
            name = excluded.name,
            category = excluded.category,
            package = COALESCE(NULLIF(excluded.package, ''), components.package),
            voltage = COALESCE(NULLIF(excluded.voltage, ''), components.voltage),
            current = COALESCE(NULLIF(excluded.current, ''), components.current),
            manufacturer = COALESCE(NULLIF(excluded.manufacturer, ''), components.manufacturer),
            purchase_date = COALESCE(NULLIF(excluded.purchase_date, ''), components.purchase_date),
            purchase_id = COALESCE(NULLIF(excluded.purchase_id, ''), components.purchase_id),
            description = COALESCE(NULLIF(excluded.description, ''), components.description),
            updated_at = excluded.updated_at
        """,
        (
            part_number,
            name,
            sheet_name,
            package,
            voltage,
            current,
            vendor,
            purchase_date,
            purchase_id,
            merged_description,
            now,
            now,
        ),
    )

    conn.execute(
        """
        INSERT INTO inventory (part_number, quantity, location, updated_at)
        VALUES (?, ?, ?, ?)
        ON CONFLICT(part_number) DO UPDATE SET
            quantity = excluded.quantity,
            location = COALESCE(NULLIF(excluded.location, ''), inventory.location),
            updated_at = excluded.updated_at
        """,
        (
            part_number,
            quantity if quantity is not None else 0,
            location,
            now,
        ),
    )

    return part_number


def _make_fallback_part_number(*, sheet_name: str, name: str, row: dict[str, Any]) -> str:
    stable_text = f"{sheet_name}|{name}|{repr(sorted(row.items()))}"
    digest = sha1(stable_text.encode("utf-8")).hexdigest()[:10]
    return f"AUTO_{digest}" if name or row else ""


def _build_sheet_column_graph(
    *,
    ontology_store: OntologyStore,
    edge_seen: set[tuple[str, str, str]],
    sheet_name: str,
    headers: list[str],
    table_name: str,
    row_count: int,
) -> tuple[str, dict[str, str]]:
    sheet_token = _safe_token(sheet_name)
    sheet_node_id = f"sheet:{sheet_token}"
    ontology_store.upsert_node(
        node_id=sheet_node_id,
        name=sheet_name,
        node_type="Sheet",
        aliases=[sheet_name],
        attributes={"table": table_name, "rows": row_count},
    )

    column_node_ids: dict[str, str] = {}
    for header in headers:
        column_token = _safe_token(header)
        column_node_id = f"column:{sheet_token}:{column_token}"
        column_node_ids[header] = column_node_id
        ontology_store.upsert_node(
            node_id=column_node_id,
            name=header,
            node_type="Column",
            aliases=[header],
            attributes={"sheet": sheet_name},
        )
        _add_edge(
            ontology_store=ontology_store,
            edge_seen=edge_seen,
            source=sheet_node_id,
            target=column_node_id,
            relation="has_column",
        )

    return sheet_node_id, column_node_ids


def _build_record_graph(
    *,
    ontology_store: OntologyStore,
    edge_seen: set[tuple[str, str, str]],
    sheet_name: str,
    sheet_node_id: str,
    row: dict[str, Any],
    row_index: int,
    column_node_ids: dict[str, str],
    mapping: SheetColumnMap,
    part_number: str | None,
) -> None:
    sheet_token = _safe_token(sheet_name)
    record_node_id = f"record:{sheet_token}:{row_index}"

    name = _to_text(row.get(mapping.name)) if mapping.name else ""
    vendor = _to_text(row.get(mapping.vendor)) if mapping.vendor else ""
    purchase_id = _to_text(row.get(mapping.purchase_id)) if mapping.purchase_id else ""

    display_name = name or part_number or purchase_id or record_node_id

    aliases = [
        value
        for value in [part_number, name, vendor, purchase_id]
        if value
    ]

    compact_attrs = {
        key: value
        for key, value in row.items()
        if value not in (None, "")
    }
    compact_attrs["sheet"] = sheet_name

    ontology_store.upsert_node(
        node_id=record_node_id,
        name=display_name,
        node_type="Record",
        aliases=aliases,
        attributes=compact_attrs,
    )

    _add_edge(
        ontology_store=ontology_store,
        edge_seen=edge_seen,
        source=sheet_node_id,
        target=record_node_id,
        relation="contains_record",
    )

    if part_number:
        component_node_id = f"component:{_safe_token(part_number)}"
        component_name = name or part_number
        ontology_store.upsert_node(
            node_id=component_node_id,
            name=component_name,
            node_type="Component",
            aliases=[part_number, component_name, vendor] if vendor else [part_number, component_name],
            attributes={
                "sheet": sheet_name,
                "part_number": part_number,
            },
        )
        _add_edge(
            ontology_store=ontology_store,
            edge_seen=edge_seen,
            source=record_node_id,
            target=component_node_id,
            relation="maps_to_component",
        )

    for header, value in row.items():
        if value in (None, ""):
            continue
        column_node_id = column_node_ids.get(header)
        if not column_node_id:
            continue
        _add_edge(
            ontology_store=ontology_store,
            edge_seen=edge_seen,
            source=record_node_id,
            target=column_node_id,
            relation="has_field",
            attributes={"value": _short_text(value)},
        )


def _add_edge(
    *,
    ontology_store: OntologyStore,
    edge_seen: set[tuple[str, str, str]],
    source: str,
    target: str,
    relation: str,
    attributes: dict[str, Any] | None = None,
) -> None:
    key = (source, target, relation)
    if key in edge_seen:
        return
    edge_seen.add(key)
    ontology_store.edges.append(
        OntologyEdge(
            source=source,
            target=target,
            relation=relation,
            attributes=attributes or {},
        )
    )


def _normalize_cell(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day).isoformat()
    if isinstance(value, str):
        return value.strip()
    return value


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _safe_token(text: str) -> str:
    cleaned = re.sub(r"[^0-9A-Za-z_\u4e00-\u9fff]+", "_", text).strip("_")
    if not cleaned:
        return f"id_{sha1(text.encode('utf-8')).hexdigest()[:8]}"
    return cleaned.lower()


def _q(identifier: str) -> str:
    return '"' + identifier.replace('"', '""') + '"'


def _to_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def _to_int(value: Any) -> int | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)

    text = str(value).strip().replace(",", "")
    if not text:
        return None

    if re.fullmatch(r"-?\d+", text):
        return int(text)
    if re.fullmatch(r"-?\d+\.\d+", text):
        return int(float(text))
    return None


def _short_text(value: Any, limit: int = 120) -> str:
    text = _to_text(value)
    if len(text) <= limit:
        return text
    return text[: limit - 3] + "..."
